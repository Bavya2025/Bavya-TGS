from rest_framework import generics, viewsets, status, serializers
from django.core.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import (
    Trip, Expense, TravelClaim, TravelAdvance, TripOdometer, Dispute, PolicyDocument, BulkActivityBatch, JobReport,
    TravelModeMaster, BookingTypeMaster, AirlineMaster, BusTypeMaster, IntercityCabVehicleMaster, TravelProviderMaster,
    LocalTravelModeMaster, LocalProviderMaster,
    StayTypeMaster, RoomTypeMaster, MealCategoryMaster, MealTypeMaster, IncidentalTypeMaster,
    CustomMasterDefinition, CustomMasterValue, MasterModule, TripTracking,
    TravelOperatorMaster, TravelClassMaster, TravelVehicleMaster, LocalSubTypeMaster
)
from .serializers import (
    TripSerializer, ExpenseSerializer, TravelClaimSerializer, TravelAdvanceSerializer,
    TripOdometerSerializer, DisputeSerializer, PolicyDocumentSerializer, BulkActivityBatchSerializer, JobReportSerializer,
    TravelModeMasterSerializer, BookingTypeMasterSerializer, AirlineMasterSerializer,
    BusTypeMasterSerializer, IntercityCabVehicleMasterSerializer, TravelProviderMasterSerializer,
    LocalTravelModeMasterSerializer, LocalProviderMasterSerializer, StayTypeMasterSerializer, RoomTypeMasterSerializer,
    MealCategoryMasterSerializer, MealTypeMasterSerializer, IncidentalTypeMasterSerializer,
    CustomMasterDefinitionSerializer, CustomMasterValueSerializer, MasterModuleSerializer,
    TripTrackingSerializer,
    LocalSubTypeMasterSerializer, TravelOperatorMasterSerializer, TravelClassMasterSerializer, TravelVehicleMasterSerializer
)
import io
import json
import pandas as pd
from django.http import HttpResponse
from rest_framework.permissions import AllowAny
from django.db.models import Q
import base64
import binascii
from api_management.utils import encrypt_key, decrypt_key
from django.utils import timezone
from rest_framework.views import APIView
from django.db.models import Sum
from django_filters.rest_framework import DjangoFilterBackend
from core.models import User
from notifications.models import Notification
from core.permissions import IsCustomAuthenticated
import requests
import datetime

def decode_id(encoded_id):
    import base64
    import binascii
    
    if not encoded_id:
        return None
        
    try:
        # Check if it's already a numeric-looking ID or doesn't look like base64
        if encoded_id.isdigit() or encoded_id.startswith('TRP-') or encoded_id.startswith('ITS-'):
            return encoded_id
            
        padding = 4 - (len(encoded_id) % 4)
        if padding != 4:
            encoded_id += '=' * padding
        
        encoded_id = encoded_id.replace('-', '+').replace('_', '/')
        
        decoded_bytes = base64.b64decode(encoded_id)
        return decoded_bytes.decode('utf-8')
    except (binascii.Error, UnicodeDecodeError, ValueError):
        return encoded_id

def _is_admin(user):
    """Checks if a user has administrative privileges."""
    user_role = (user.role.name.lower() if user.role else '')
    return user_role in ['admin', 'it-admin', 'superuser']

def _is_finance_head(user):
    """Checks if a user is the Finance Head."""
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'head' in dept and 'finance' in dept or 'head' in desig and 'finance' in desig or 'cfo' in user_role

def _is_finance_executive(user):
    """Checks if a user is a Finance Executive."""
    if _is_finance_head(user): return False
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'finance' in user_role or 'finance' in dept or 'finance' in desig

def _is_hr(user):
    """Checks if a user is an HR user."""
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'hr' in user_role or 'hr' in dept or 'hr' in desig

def _get_finance_users():
    """Returns a list of users who should be treated as Finance."""
    all_users = User.objects.filter(is_active=True).select_related('role')
    return [u for u in all_users if _is_finance_executive(u) or _is_finance_head(u)]

def _get_hr_users():
    """Returns a list of users who should be treated as HR."""
    all_users = User.objects.filter(is_active=True).select_related('role')
    return [u for u in all_users if _is_hr(u)]

def get_finance_head(user, exclude_user=None):
    """Finds a finance approver (Head of Finance)."""
    all_finance = _get_finance_users()
    heads = [u for u in all_finance if _is_finance_head(u)]
    if exclude_user:
        heads = [u for u in heads if u.id != exclude_user.id]
    
    if heads:
        return heads[0]
    # Fallback to any finance user if no head found
    return all_finance[0] if all_finance else None

def get_finance_executive(user, exclude_user=None):
    """Finds a finance executive."""
    all_finance = _get_finance_users()
    execs = [u for u in all_finance if _is_finance_executive(u)]
    if exclude_user:
         execs = [u for u in execs if u.id != exclude_user.id]
         
    if execs:
        return execs[0]
    return all_finance[0] if all_finance else None

def get_hr_head(user):
    """Finds an HR approver (Head of HR)."""
    all_hr = _get_hr_users()
    # Try local HR first
    local_hr = [u for u in all_hr if u.base_location == user.base_location]
    if local_hr:
        return local_hr[0]
    
    return all_hr[0] if all_hr else None

    local_heads = [u for u in heads if u.base_location == user.base_location]
    return local_heads[0] if local_heads else (heads[0] if heads else None)

def notify_hr(title, message):
    """Notify all users with HR role."""
    hr_users = User.objects.filter(role__name__icontains='hr')
    for hr in hr_users:
        Notification.objects.create(
            user=hr,
            title=title,
            message=message,
            type='info'
        )


def update_trip_lifecycle(trip, title, description):
    """Helper to append events to the Trip's JSON lifecycle field."""
    if not trip:
        return
    event = {
        "title": title,
        "status": "completed",
        "date": timezone.now().strftime("%b %d, %Y"),
        "description": description
    }
    events = trip.lifecycle_events
    if not isinstance(events, list):
        events = []
    
    # Avoid duplicate events with same title
    if not any(e.get('title') == title for e in events):
        events.append(event)
        trip.lifecycle_events = events
        trip.save(update_fields=['lifecycle_events'])






def resolve_approver(user, members_data=None):
    """Helper to resolve the first approver in the management hierarchy."""
    def is_admin(u):
        if not u or not u.role:
            return False
        return u.role.name.lower() in ['admin', 'it-admin', 'superuser', 'it admin', 'system administrator']
    
    reporting_manager = user.reporting_manager
    senior_manager = user.senior_manager
    hod_director = user.hod_director
    
    current_approver = reporting_manager if not is_admin(reporting_manager) else None
    h_level = 1
    
    if not current_approver:
        current_approver = senior_manager if not is_admin(senior_manager) else None
        h_level = 2
    
    if not current_approver:
        current_approver = hod_director if not is_admin(hod_director) else None
        h_level = 3
        
    if not current_approver:
        # Fallback to members' managers if applicable
        potential_managers = []
        if members_data:
            import re
            for m_str in members_data:
                match = re.search(r'\((.*?)\)', m_str)
                if match:
                    member_id = match.group(1)
                    member_user = User._get_or_create_shell_user(member_id)
                    manager = member_user.reporting_manager if member_user else None
                    if manager and not is_admin(manager):
                        potential_managers.append(manager)
            
        if potential_managers:
            potential_managers.sort(key=lambda m: getattr(m, 'level_rank', 99))
            current_approver = potential_managers[0]
            h_level = 1
        else:
            current_approver = get_hr_head(user)
            h_level = 1
            
    return current_approver, h_level, reporting_manager, senior_manager, hod_director

class TripListCreateView(generics.ListCreateAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()
            
        all_trips = self.request.query_params.get('all') == 'true'
        user_role = user.role.name.lower() if user.role else ''
        
        search_query = self.request.query_params.get('search', None)
        if search_query:
             # Support potential Base64 encoding if used by frontend
             try:
                import base64
                if len(search_query) > 4 and ('-' in search_query or '_' in search_query or search_query.isalpha()):
                    # Likely search text
                    pass
                else:
                    padding = len(search_query) % 4
                    if padding: search_query += '=' * (4 - padding)
                    search_query = base64.b64decode(search_query).decode('utf-8')
             except: pass
             
             from django.db.models import Q
             queryset = Trip.objects.filter(
                Q(trip_id__icontains=search_query) |
                Q(destination__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(trip_leader__icontains=search_query)
             )
             if not all_trips:
                 queryset = queryset.filter(user=user, consider_as_local=False)
             return queryset.order_by('-created_at')

        if all_trips:
            if user_role in ['admin', 'guesthousemanager', 'finance', 'cfo']:
                return Trip.objects.all().order_by('-created_at')
            
            from django.db.models import Q
            return Trip.objects.filter(
                Q(user=user) | 
                Q(current_approver=user)
            ).distinct().order_by('-created_at')
            
        return Trip.objects.filter(user=user, consider_as_local=False).order_by('-created_at')

    def perform_create(self, serializer, is_local=False):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            from rest_framework.exceptions import AuthenticationFailed
            raise AuthenticationFailed("Authentication required.")
        
        # Admin / Superuser skip approvals
        user_role = user.role.name.lower() if user.role else ''
        if user_role in ['admin', 'superuser', 'it-admin']:
            trip = serializer.save(
                user=user,
                status='Approved',
                current_approver=None,
                hierarchy_level=0,
                consider_as_local=is_local
            )
            label = "Travel" if is_local else "Trip"
            update_trip_lifecycle(trip, "Auto-Approved", f"{label} request auto-approved for Administrator.")
            return

        members_data = serializer.validated_data.get('members', [])
        current_approver, h_level, rm, sm, hod = resolve_approver(user, members_data)
        
        try:
            trip = serializer.save(
                user=user,
                status='Pending',
                current_approver=current_approver,
                hierarchy_level=h_level,
                consider_as_local=is_local,
                # Snapshots
                user_name=user.name,
                user_designation=user.designation,
                user_department=user.department,
                reporting_manager_name=rm.name if rm else None,
                senior_manager_name=sm.name if sm else None,
                hod_director_name=hod.name if hod else None
            )
        except Exception as e:
            # convert DB errors to validation error so frontend sees message
            # also log full traceback on server for diagnostics
            import traceback
            traceback.print_exc()
            from rest_framework.exceptions import ValidationError
            msg = str(e)
            print("ERROR IN TRIP CREATION:", msg)
            raise ValidationError({"detail": msg})

        label = "Travel" if is_local else "Trip"
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title=f"New {label} Request",
                message=f"{user.name} has submitted a new {label.lower()} request to {trip.destination} (ID: {trip.trip_id}).",
                type='info'
            )
            
            # NOTIFY USER: Sent for approval
            Notification.objects.create(
                user=user,
                title=f"{label} Request Sent",
                message=f"Your {label.lower()} request {trip.trip_id} has been created and sent to {current_approver.name} for approval.",
                type='success'
            )
        else:
            # Auto-approved or no approver
            Notification.objects.create(
                user=user,
                title=f"{label} Request Created",
                message=f"Your {label.lower()} request {trip.trip_id} has been successfully created.",
                type='success'
            )
        
        if trip.accommodation_requests and any('Room' in r for r in trip.accommodation_requests):
            gh_managers = User.objects.filter(role__name='GuestHouseManager', is_active=True)
            for manager in gh_managers:
                Notification.objects.create(
                    user=manager,
                    title="Room Request Received",
                    message=f"{user.name} has requested a room for {label.lower()} {trip.trip_id}.",
                    type='info'
                )

        notify_hr(f"New {label} Request", f"{user.name} has raised a {label.lower()} request to {trip.destination} (ID: {trip.trip_id}).")

class TravelListCreateView(TripListCreateView):
    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user: return Trip.objects.none()
        
        queryset = Trip.objects.filter(user=user, consider_as_local=True)
        search_query = self.request.query_params.get('search', None)
        if search_query:
             from django.db.models import Q
             queryset = queryset.filter(
                Q(trip_id__icontains=search_query) |
                Q(destination__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(trip_leader__icontains=search_query)
             )
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        super().perform_create(serializer, is_local=True)

class TripBookingSearchView(generics.ListAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated] 

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()

        # Users only search their own trips in this view
        queryset = Trip.objects.filter(user=user)

        search_query = self.request.query_params.get('search', None)
        if search_query:
            try:
                import base64
                padding = len(search_query) % 4
                if padding: 
                    search_query += '=' * (4 - padding)
                search_query = base64.b64decode(search_query).decode('utf-8')
            except Exception as e:
                print(f"Decoding error: {e}")
                pass

            queryset = queryset.filter(
                Q(trip_id__istartswith=search_query) | 
                Q(purpose__istartswith=search_query) | 
                Q(source__istartswith=search_query) | 
                Q(destination__istartswith=search_query) | 
                Q(trip_leader__istartswith=search_query)
            )
        
        return queryset

class TripDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Trip.objects.all()
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]
    lookup_field = 'trip_id'

    def get_object(self):
        if 'trip_id' in self.kwargs:
            self.kwargs['trip_id'] = decode_id(self.kwargs['trip_id'])
        
        obj = super().get_object()
        user = getattr(self.request, 'custom_user', None)
        
        if not user:
            self.permission_denied(self.request, message="Not authenticated")
            
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = user_role in ['finance', 'cfo']
        is_gh_manager = user_role == 'guesthousemanager'
        
        # Authorization check: Owner, Managers in the hierarchy, Current Approver, Finance, Guest House Manager, or Admin
        is_owner = (obj.user == user)
        # Check if the current user is any of the requester's managers or the current approver
        is_manager = user in [obj.user.reporting_manager, obj.user.senior_manager, obj.user.hod_director, obj.current_approver]
        
        if not (is_owner or is_manager or is_admin or is_finance or is_gh_manager):
             self.permission_denied(self.request, message="Not authorized to view this trip details")
             
        return obj

class TripTrackingView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.get called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        # Verify trip exists and user has access
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            print(f"DEBUG: Trip {real_trip_id} not found")
            return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Basic access check: requester or manager or finance or admin
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: Requester: {user.employee_id if user else 'Anonymous'}")
        
        # ... existing logic ...
        is_owner = (trip.user == user)
        is_manager = False
        if user:
            is_manager = user in [trip.user.reporting_manager, trip.user.senior_manager, trip.user.hod_director, trip.current_approver]
        
        user_role = user.role.name.lower() if user and user.role else ''
        is_privileged = user_role in ['admin', 'finance', 'cfo', 'guesthousemanager']

        if not (is_owner or is_manager or is_privileged):
            print(f"DEBUG: Unauthorized access attempt to trip {real_trip_id}")
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        tracking_data = TripTracking.objects.filter(trip=trip).order_by('timestamp')
        print(f"DEBUG: Returning {tracking_data.count()} points")
        serializer = TripTrackingSerializer(tracking_data, many=True)
        return Response(serializer.data)

    def post(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.post called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            print(f"DEBUG: Trip {real_trip_id} not found for POST")
            return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Only trip owner can post tracking points
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: POST Requester: {user.employee_id if user else 'Anonymous'}")
        
        if not user or trip.user != user:
            print(f"DEBUG: POST Unauthorized for user {user.employee_id if user else 'None'}")
            return Response({"error": "Only trip owner can submit tracking data"}, status=status.HTTP_403_FORBIDDEN)

        data = request.data.copy()
        data['trip'] = trip.trip_id
        
        serializer = TripTrackingSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            print("DEBUG: Tracking point saved successfully")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        print(f"DEBUG: Serializer errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ApprovalCountView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"total": 0, "trips": 0, "advances": 0, "claims": 0, "batches": 0})
        
        is_admin = _is_admin(user)
        is_finance_head = _is_finance_head(user)
        is_finance_exec = _is_finance_executive(user)
        is_hr = _is_hr(user)
        
        trip_count = 0
        advance_count = 0
        claim_count = 0
        batch_count = 0

        pending_money_statuses = ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE']
        finance_pending = pending_money_statuses + ['PENDING_HEAD']

        if is_admin:
            trip_count = Trip.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved']).count()
            advance_count = TravelAdvance.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending).count()
            claim_count = TravelClaim.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending).count()
            batch_count = BulkActivityBatch.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved']).count()
        elif is_finance_head:
            advance_count = TravelAdvance.objects.filter(status='PENDING_HEAD').count()
            claim_count = TravelClaim.objects.filter(status='PENDING_HEAD').count()
        elif is_finance_exec:
            advance_count = TravelAdvance.objects.filter(status__in=pending_money_statuses).count()
            claim_count = TravelClaim.objects.filter(status__in=pending_money_statuses).count()
        elif is_hr:
            trip_count = Trip.objects.filter(status='Manager Approved').count()
            advance_count = TravelAdvance.objects.filter(status='Manager Approved').count()
            claim_count = TravelClaim.objects.filter(status='Manager Approved').count()
            batch_count = BulkActivityBatch.objects.filter(status='Manager Approved').count()
        else:
            # Manager/Approver
            trip_count = Trip.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded']).count()
            advance_count = TravelAdvance.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded']).count()
            claim_count = TravelClaim.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded']).count()
            batch_count = BulkActivityBatch.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded']).count()
            
        return Response({
            "total": trip_count + advance_count + claim_count + batch_count,
            "trips": trip_count,
            "advances": advance_count,
            "claims": claim_count,
            "batches": batch_count
        })


class ApprovalsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = _is_finance_executive(user) or _is_finance_head(user)
        is_hr = _is_hr(user)
        is_finance_head = _is_finance_head(user)
        
        tab = request.query_params.get('tab', 'pending')
        type_filter = request.query_params.get('type', 'all') 
        
        trips = Trip.objects.none()
        advances = TravelAdvance.objects.none()
        claims = TravelClaim.objects.none()
        batches = BulkActivityBatch.objects.none()
        disputes = Dispute.objects.none()

        if tab == 'history':
            from core.models import AuditLog
            # Include 'UPDATE' to capture older approvals or edits made by managers
            involved_logs = AuditLog.objects.filter(user=user, action__in=['APPROVE', 'FORWARD', 'REJECT', 'UPDATE'])
            
            trip_pks = involved_logs.filter(model_name='Trip').values_list('object_id', flat=True)
            advance_pks_raw = involved_logs.filter(model_name='TravelAdvance').values_list('object_id', flat=True)
            claim_pks_raw = involved_logs.filter(model_name='TravelClaim').values_list('object_id', flat=True)
            
            # Convert string IDs to integers for numeric primary keys
            advance_pks = [int(pk) for pk in advance_pks_raw if pk and pk.isdigit()]
            claim_pks = [int(pk) for pk in claim_pks_raw if pk and pk.isdigit()]
            
            batch_pks_raw = involved_logs.filter(model_name='BulkActivityBatch').values_list('object_id', flat=True)
            batch_pks = [int(pk) for pk in batch_pks_raw if pk and pk.isdigit()]

            # --- EXTENDED HISTORY: Also include user's own finished requests ---
            history_statuses = ['Approved', 'Rejected', 'Resolved', 'Paid', 'HR Approved', 'Manager Approved', 'COMPLETED', 'Settled']
            
            trips = Trip.objects.filter(Q(trip_id__in=trip_pks) | Q(user=user, status__in=history_statuses))
            advances = TravelAdvance.objects.filter(Q(id__in=advance_pks) | Q(trip__user=user, status__in=history_statuses))
            claims = TravelClaim.objects.filter(Q(id__in=claim_pks) | Q(trip__user=user, status__in=history_statuses))
            batches = BulkActivityBatch.objects.filter(Q(id__in=batch_pks) | Q(user=user, status__in=history_statuses))
            
            # Admins see everything in history
            if is_admin:
                trips = Trip.objects.filter(status__in=history_statuses)
                advances = TravelAdvance.objects.filter(status__in=history_statuses)
                claims = TravelClaim.objects.filter(status__in=history_statuses)
                batches = BulkActivityBatch.objects.filter(status__in=history_statuses)
        else:
            # Pending Tab
            if is_admin:
                finance_pending = ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'REJECTED_BY_HEAD']
                trips = Trip.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                advances = TravelAdvance.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                claims = TravelClaim.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                batches = BulkActivityBatch.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'])
            elif is_finance:
                if is_finance_head:
                    advances = TravelAdvance.objects.filter(status='PENDING_HEAD')
                    claims = TravelClaim.objects.filter(status='PENDING_HEAD')
                else:
                    pending_money_statuses = ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE']
                    advances = TravelAdvance.objects.filter(status__in=pending_money_statuses)
                    claims = TravelClaim.objects.filter(status__in=pending_money_statuses)
                    # Finance usually doesn't approve tour plans unless they have costing
                    batches = BulkActivityBatch.objects.none()
                
                # Filter by project/dept if needed
                finance_dept = user.department
                if finance_dept and finance_dept.lower() not in ['finance', 'finance department', 'accounts', 'finance head dept', 'finance executive dept']:
                    advances = advances.filter(trip__project_code__istartswith=finance_dept)
                    claims = claims.filter(trip__project_code__istartswith=finance_dept)
            elif is_hr:
                # HR verification stage
                trips = Trip.objects.filter(status='Manager Approved')
                advances = TravelAdvance.objects.filter(status='Manager Approved')
                claims = TravelClaim.objects.filter(status='Manager Approved')
                batches = BulkActivityBatch.objects.filter(status='Manager Approved')
            else:
                # Regular hierarchy
                trips = Trip.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
                advances = TravelAdvance.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
                claims = TravelClaim.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
                batches = BulkActivityBatch.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
        
        tasks = []
        # Support filtering by type if specified
        if type_filter in ['all', 'trip']:
            for t in trips.order_by('-created_at'):
                tasks.append({
                    "id": f"TRIP-{t.trip_id}", "db_id": t.trip_id, "type": "Trip",
                    "requester": t.user.name if t.user else "Unknown", "purpose": t.purpose,
                    "status": t.status, "date": t.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": t.hierarchy_level,
                    "current_approver_name": t.current_approver.name if t.current_approver else (t.status if t.status in ['Approved', 'Rejected'] else "N/A"),
                    "trip_id": t.trip_id,
                    "is_local": t.consider_as_local,
                    "cost": t.cost_estimate,
                    "details": {
                        "source": t.source, "destination": t.destination, 
                        "start_date": t.start_date.strftime("%b %d, %Y"),
                        "end_date": t.end_date.strftime("%b %d, %Y"),
                        "travel_mode": t.travel_mode,
                        "composition": t.composition,
                        "vehicle_type": t.vehicle_type,
                        "purpose": t.purpose,
                        "project_code": t.project_code,
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in t.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(t.odometer_details.start_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(t.odometer_details.start_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_image else None,
                            "end_reading": str(t.odometer_details.end_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(t.odometer_details.end_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_image else None,
                        } if hasattr(t, 'odometer_details') else None
                    }
                })
            
        if type_filter in ['all', 'advance']:
            for a in advances.order_by('-created_at'):
                tasks.append({
                    "id": f"ADV-{a.id}", "db_id": a.id, "type": "Money Top-up / Advance",
                    "requester": a.trip.user.name if a.trip.user else "Unknown",
                    "purpose": f"Advance: {a.purpose}", "cost": f"₹{a.requested_amount}",
                    "status": a.status, "date": a.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": a.hierarchy_level,
                    "current_approver_name": a.current_approver.name if a.current_approver else (a.status if a.status in ['Approved', 'Rejected'] else "N/A"),
                    "trip_id": a.trip.trip_id,
                    "is_local": a.trip.consider_as_local,
                    "details": {
                        "source": a.trip.source,
                        "destination": a.trip.destination,
                        "requested_amount": str(a.requested_amount),
                        "hr_approved_amount": str(a.hr_approved_amount or 0),
                        "hr_remarks": a.hr_remarks or "",
                        "executive_approved_amount": str(a.executive_approved_amount),
                        "reason": a.purpose,
                        "trip_destination": a.trip.destination,
                        "trip_id": a.trip.trip_id,
                        "start_date": a.trip.start_date.strftime("%b %d, %Y") if a.trip.start_date else "N/A",
                        "end_date": a.trip.end_date.strftime("%b %d, %Y") if a.trip.end_date else "N/A",
                    }
                })

        if type_filter in ['all', 'expense', 'mileage']:
            for c in claims.order_by('-created_at'):
                tasks.append({
                    "id": f"CLAIM-{c.id}", "db_id": c.id, "type": "Expense Claim",
                    "requester": c.trip.user.name if c.trip.user else "Unknown",
                    "purpose": f"Claim for {c.trip.destination}", "cost": f"₹{c.total_amount}",
                    "status": c.status, "date": c.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": c.hierarchy_level,
                    "current_approver_name": c.current_approver.name if c.current_approver else (c.status if c.status in ['Approved', 'Rejected'] else "N/A"),
                    "trip_id": c.trip.trip_id,
                    "is_local": c.trip.consider_as_local,
                    "details": {
                        "source": c.trip.source,
                        "destination": c.trip.destination,
                        "total_amount": str(c.total_amount),
                        "requested_amount": str(c.total_amount),
                        "approved_amount": str(c.approved_amount),
                        "hr_approved_amount": str(c.hr_approved_amount or 0),
                        "hr_remarks": getattr(c, "hr_remarks", ""),
                        "executive_approved_amount": str(c.executive_approved_amount),
                        "trip_id": c.trip.trip_id,
                        "start_date": c.trip.start_date.strftime("%b %d, %Y") if c.trip.start_date else "N/A",
                        "end_date": c.trip.end_date.strftime("%b %d, %Y") if c.trip.end_date else "N/A",
                        "expenses": [
                            {
                                "id": e.id,
                                "category": e.category,
                                "date": e.date.strftime("%b %d, %Y"),
                                "amount": str(e.amount),
                                "description": e.description,
                                "status": e.status,
                                "receipt_image": decrypt_key(e.receipt_image) if e.receipt_image else "",
                                "rm_remarks": e.rm_remarks or "",
                                "hr_remarks": e.hr_remarks or "",
                                "finance_remarks": e.finance_remarks or ""
                            } for e in c.trip.expenses.all()
                        ],
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in c.trip.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(c.trip.odometer_details.start_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(c.trip.odometer_details.start_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_image else None,
                            "end_reading": str(c.trip.odometer_details.end_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(c.trip.odometer_details.end_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_image else None,
                        } if hasattr(c.trip, 'odometer_details') else None
                    }
                })

        if type_filter in ['all', 'batch']:
            for b in batches.order_by('-created_at'):
                tasks.append({
                    "id": f"BATCH-{b.id}", "db_id": b.id, "type": "Monthly Tour Plan",
                    "requester": b.user.name if b.user else "Unknown",
                    "user_name": b.user.name if b.user else "Unknown",
                    "purpose": f"Tour Plan: {b.file_name}", "status": b.status,
                    "date": b.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": b.hierarchy_level,
                    "file_name": b.file_name,
                    "data_json": b.data_json,
                    "current_approver_name": b.current_approver.name if b.current_approver else (b.status if b.status in ['Approved', 'Rejected'] else "N/A"),
                    "trip_id": b.trip.trip_id if b.trip else "N/A",
                    "is_local": True,
                    "details": {
                        "file_name": b.file_name,
                        "entry_count": len(b.data_json) if isinstance(b.data_json, list) else 0,
                        "data": b.data_json,
                        "remarks": b.remarks
                    }
                })

        for d in disputes.order_by('-created_at'):
            tasks.append({
                "id": f"DISPUTE-{d.id}", "db_id": d.id, "type": "Dispute",
                "requester": d.raised_by.name if d.raised_by else "Unknown",
                "purpose": f"Dispute: {d.category}", "status": d.status,
                "date": d.created_at.strftime("%b %d, %Y")
            })

        return Response(tasks)

    def post(self, request):
        user = getattr(request, 'custom_user', None)
        task_id = request.data.get('id')
        action = request.data.get('action') 
        
        if not task_id or not action:
            return Response({"error": "ID and Action required"}, status=400)
            
        if action == 'UpdateItem':
            item_id = request.data.get('item_id')
            item_status = request.data.get('item_status')
            remarks = request.data.get('remarks', '')
            if item_id:
                from .models import Expense
                expense = Expense.objects.filter(id=item_id).first()
                if expense:
                    update_fields = {'status': item_status}
                    # Determine which remarks field to update based on user role
                    user_role = user.role.name.upper() if user and user.role else ""
                    if "FINANCE" in user_role:
                        update_fields['finance_remarks'] = remarks
                    elif "HR" in user_role:
                        update_fields['hr_remarks'] = remarks
                    else:
                        update_fields['rm_remarks'] = remarks
                    
                    Expense.objects.filter(id=item_id).update(**update_fields)
                    return Response({"message": f"Item {item_id} set to {item_status} with remarks"})
            return Response({"error": "Item ID required or not found"}, status=400)

        try:
            if task_id.startswith('TRIP-'):
                obj = Trip.objects.get(trip_id=task_id.replace('TRIP-', ''))
            elif task_id.startswith('ADV-'):
                obj = TravelAdvance.objects.get(id=task_id.replace('ADV-', ''))
            elif task_id.startswith('CLAIM-'):
                obj = TravelClaim.objects.get(id=task_id.replace('CLAIM-', ''))
            elif task_id.startswith('DISPUTE-'):
                obj = Dispute.objects.get(id=task_id.replace('DISPUTE-', ''))
            else:
                return Response({"error": "Invalid task ID"}, status=400)
            
            self._handle_workflow(obj, action, user, request.data)
            return Response({"message": f"Task processed successfully: {action}"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=400)

    def _handle_workflow(self, obj, action, user, data=None):
        """Generic multi-level workflow handler"""
        # Determine requester 
        requester = obj.user if hasattr(obj, 'user') else obj.trip.user
        request_type = "Trip" if isinstance(obj, Trip) else ("Advance" if isinstance(obj, TravelAdvance) else "Expense Claim")
        
        # Security Check: Determine roles and check authorization early
        user_role = user.role.name.lower() if user.role else ''
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_hr = _is_hr(user)
        is_finance_exec = _is_finance_executive(user)
        is_finance_head = _is_finance_head(user)
        is_finance = is_finance_exec or is_finance_head

        if not is_admin:
            authorized = False
            # 1. Primary approver check
            if obj.current_approver == user:
                authorized = True
            # 2. Functional role check (allow acting on specific statuses even if not the explicit current_approver)
            elif is_hr and obj.status == 'Manager Approved':
                authorized = True
            elif is_finance_exec and obj.status in ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved', 'Under Process']:
                authorized = True
            elif is_finance_head and obj.status in ['PENDING_HEAD', 'PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved', 'Under Process']:
                authorized = True
            # 3. Finance-specific actions check
            elif is_finance and action in ['Transfer', 'Pay', 'UnderProcess', 'RejectByFinance']:
                authorized = True

            if not authorized:
                raise Exception("You are not authorized to perform this action on this request.")

        from core.models import AuditLog
        if action == 'Reject':
            obj.status = 'Rejected'
            obj.current_approver = None
            obj.save()
            
            AuditLog.objects.create(
                user=user, action='REJECT', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj),
                details={'reason': data.get('remarks') if data else ''}
            )

        if action == 'Forward':
            # This is now mostly handled by 'Approve' automatic progression, 
            # but we keep it for manual overrides if needed by admins.
            mgr_l2 = getattr(obj, 'senior_manager', None)
            mgr_l3 = getattr(obj, 'hod_director', None)
            
            next_approver = None
            next_level = obj.hierarchy_level
            
            if obj.hierarchy_level == 1:
                next_approver = mgr_l2 or mgr_l3
                next_level = 2 if mgr_l2 else 3
            elif obj.hierarchy_level == 2:
                next_approver = mgr_l3
                next_level = 3
            
            if not next_approver:
                raise Exception("There is no higher-level manager to forward this request to.")

            obj.current_approver = next_approver
            obj.hierarchy_level = next_level
            obj.save()
            
            AuditLog.objects.create(
                user=user, action='FORWARD', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj),
                details={'to': str(next_approver)}
            )
            return Response({"message": "Forwarded successfully"})

        if action == 'Approve':
            is_hr = _is_hr(user)
            is_finance = _is_finance_executive(user) or _is_finance_head(user)
            from core.models import AuditLog
            AuditLog.objects.create(
                user=user, action='APPROVE', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj)
            )
            
            # --- STAGE 1: Management Hierarchy ---
            if not is_hr and not is_finance:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                if trip:
                    level = getattr(obj, 'hierarchy_level', 1)
                    
                    # If this is a claim, calculate the approved total based on item statuses
                    if isinstance(obj, TravelClaim):
                        from django.db.models import Sum
                        # Re-calculate approved total from items not marked as Rejected
                        # We assume anything not 'Rejected' is 'Approved' or 'Pending' (which defaults to OK in this step)
                        # Actually, better to strictly count only 'Approved' or 'Pending'
                        approved_sum = obj.trip.expenses.exclude(status='Rejected').aggregate(s=Sum('amount'))['s'] or 0
                        obj.approved_amount = approved_sum
                        obj.save()
                        
                        rejected_items = obj.trip.expenses.filter(status='Rejected')
                        rej_msg = f" with {rejected_items.count()} items rejected" if rejected_items.exists() else ""
                        update_trip_lifecycle(trip, f"Level {level} Approval", f"Approved by {user.name}{rej_msg}. Net Approved: ₹{approved_sum}.")
                    else:
                        update_trip_lifecycle(trip, f"Level {level} Approval", f"Approved by {user.name}.")
                
                next_approver = None
                
                # Try explicit levels first
                if obj.hierarchy_level == 1:
                    next_approver = getattr(obj, 'senior_manager', None) or getattr(obj, 'hod_director', None)
                elif obj.hierarchy_level == 2:
                    next_approver = getattr(obj, 'hod_director', None)
                
                # DYNAMIC FALLBACK: If no explicit level but current user has a manager
                if not next_approver:
                    # Use dynamic property to find manager
                    potential_manager = user.reporting_manager
                    
                    # Ensure we don't route back to requester or to a non-existent manager
                    if potential_manager and potential_manager != user and potential_manager != requester:
                        next_approver = potential_manager
                
                if next_approver:
                    # Move to next manager in chain
                    obj.current_approver = next_approver
                    obj.hierarchy_level += 1
                    obj.save()
                    
                    Notification.objects.create(
                        user=next_approver,
                        title=f"Pending {request_type} Approval [{obj.trip_id}]",
                        message=f"{requester.name}'s {request_type} {obj.trip_id} requires your review (Forwarded from {user.name}).",
                        type='info'
                    )

                    Notification.objects.create(
                        user=requester,
                        title=f"Approved by {user.name} [{obj.trip_id}]",
                        message=f"Your {request_type} {obj.trip_id} has been approved by {user.name} and forwarded to {next_approver.name} for review.",
                        type='success'
                    )
                else:
                    # End of Management Chain -> Move to HR Approval
                    hr_head = get_hr_head(requester)
                    obj.status = 'Manager Approved'
                    obj.current_approver = hr_head
                    obj.save()
                    
                    Notification.objects.create(
                        user=requester,
                        title=f"Management Approved [{obj.trip_id}]",
                        message=f"Your {request_type} {obj.trip_id} has been approved by management and sent to HR for verification.",
                        type='success'
                    )

                    # Notify Guest House Managers if room request exists and this is a Trip
                    if isinstance(obj, Trip) and obj.accommodation_requests and 'Request for Room' in obj.accommodation_requests:
                        gh_managers = User.objects.filter(role__name='GuestHouseManager')
                        for ghm in gh_managers:
                            Notification.objects.create(
                                user=ghm,
                                title="Pending Room Request",
                                message=f"{requester.name}'s trip to {obj.destination} is management-approved. Room booking may be initiated.",
                                type='info'
                            )
                    
                    if hr_head:
                        Notification.objects.create(
                            user=hr_head,
                            title=f"HR Verification Required [{obj.trip_id}]",
                            message=f"{requester.name}'s {request_type} {obj.trip_id} is management-approved and awaits your verification.",
                            type='info'
                        )
                return Response({"message": "Sent to HR for verification"})

            # --- STAGE 2: HR Approval ---
            elif is_hr:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                if trip:
                    update_trip_lifecycle(trip, "Ticket Booking", f"HR ({user.name}) has verified travel logistics.")
                
                if isinstance(obj, Trip):
                    # --- TRIP ENDS AT HR ---
                    obj.status = 'Approved'
                    obj.current_approver = None
                    obj.save()
                    
                    Notification.objects.create(
                        user=requester,
                        title=f"Trip Approved [{obj.trip_id}]",
                        message=f"Your Trip {obj.trip_id} to {obj.destination} has been final-approved by HR.",
                        type='success'
                    )

                    # Notify Guest House Managers if room request exists
                    if obj.accommodation_requests and 'Request for Room' in obj.accommodation_requests:
                        gh_managers = User.objects.filter(role__name='GuestHouseManager')
                        for ghm in gh_managers:
                            Notification.objects.create(
                                user=ghm,
                                title="Room Request Ready",
                                message=f"Trip {obj.trip_id} to {obj.destination} has been approved. Room booking is now required.",
                                type='info'
                            )
                else:
                    # --- MONEY REQUESTS MOVE TO FINANCE EXECUTIVE ---
                    # Capture HR specific fields
                    hr_amt = data.get('hr_approved_amount')
                    hr_rem = data.get('hr_remarks') or data.get('remarks')
                    
                    if hasattr(obj, 'hr_approved_amount'):
                        obj.hr_approved_amount = hr_amt if hr_amt is not None else getattr(obj, 'requested_amount', getattr(obj, 'total_amount', 0))
                    if hasattr(obj, 'hr_remarks'):
                        obj.hr_remarks = hr_rem
                    
                    finance_exec = get_finance_executive(requester, exclude_user=user)
                    obj.status = 'PENDING_EXECUTIVE'
                    obj.current_approver = finance_exec
                    obj.save()
                    
                    Notification.objects.create(
                        user=requester,
                        title=f"HR Verified [{obj.trip_id}]",
                        message=f"Your {request_type} {obj.trip_id} has been verified by HR and sent for finance review.",
                        type='success'
                    )
                    
                    if finance_exec:
                        Notification.objects.create(
                            user=finance_exec,
                            title=f"Finance Verification Required [{obj.trip_id}]",
                            message=f"{requester.name}'s {request_type} {obj.trip_id} is HR-verified and awaits your verification.",
                            type='info'
                        )
                return Response({"message": "HR recommendation processed"})
            # --- STAGE 3: Finance Approval ---
            elif is_finance:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                
                # Case A: Finance Executive Verification (Finance Executive 1)
                if _is_finance_executive(user) and obj.status in ['PENDING_EXECUTIVE', 'HR Approved', 'REJECTED_BY_HEAD', 'Approved', 'Under Process']:
                    exec_amount = data.get('approved_amount') or data.get('executive_approved_amount')
                    if exec_amount is None:
                        exec_amount = getattr(obj, 'requested_amount', getattr(obj, 'total_amount', 0))
                    
                    obj.executive_approved_amount = exec_amount
                    obj.sent_by_executive = user
                    obj.status = 'PENDING_HEAD'
                    obj.current_approver = get_finance_head(requester, exclude_user=user)
                    obj.save()
                    
                    if trip:
                        update_trip_lifecycle(trip, "Finance Verified", f"Verified by Executive {user.name}. Amount recommended: ₹{exec_amount}.")
                    
                    Notification.objects.create(
                        user=obj.current_approver,
                        title=f"Finance Authorization Required [{obj.trip_id}]",
                        message=f"{requester.name}'s request {obj.trip_id} verified by executive and awaits your authorization.",
                        type='info'
                    )
                    return Response({"message": "Verified and sent to Head"})

                # Case B: Finance Head Authorization (from PENDING_HEAD)
                if _is_finance_head(user) and obj.status == 'PENDING_HEAD':
                    if action == 'Approve':
                        obj.status = 'PENDING_FINAL_RELEASE'
                        obj.head_action = 'Approved'
                        
                        # Forward to OTHER finance executive (Finance Executive 2)
                        all_fin_users = _get_finance_users()
                        other_execs = [u for u in all_fin_users if not _is_finance_head(u) and u != obj.sent_by_executive]
                        final_exec = other_execs[0] if other_execs else obj.sent_by_executive
                        
                        obj.current_approver = final_exec
                        obj.final_executive = final_exec
                        obj.save()
                        
                        if trip:
                            update_trip_lifecycle(trip, "Finance Authorized", f"Authorized by Head {user.name}. Sent to {final_exec.name if final_exec else 'Executive'} for payout.")
                    
                    elif action == 'Reject':
                        obj.status = 'REJECTED_BY_HEAD'
                        obj.head_action = 'Rejected'
                        obj.current_approver = obj.sent_by_executive
                        obj.save()
                        
                        if trip:
                            update_trip_lifecycle(trip, "Head Rejected", f"Rejected by Finance Head {user.name}. Returned to Executive for revision.")
                    return Response({"message": "Head action processed"})

                # Case C: Transfer action handled below in Finance Specific Actions section
                pass

                return Response({"message": "No specific finance action matched"}, status=200)

        # Finance Specific Actions
        if action == 'UnderProcess':
            obj.status = 'Under Process'
            obj.save()
            Notification.objects.create(
                user=requester,
                title=f"Finance: Under Process [{obj.trip_id}]",
                message=f"Your {request_type} {obj.trip_id} is under process by Finance Team.",
                type='info'
            )
            return Response({"message": "Under process updated"})

        if action in ['Transfer', 'Pay']:
            # Security: Only finance can perform transfers, and usually only for finalized requests
            if not _is_finance_executive(user) and not _is_finance_head(user):
                raise PermissionDenied("Only Finance can process transfers.")
            
            # Additional check: If it has a final_executive assigned, only they should pay
            if hasattr(obj, 'final_executive') and obj.final_executive and obj.final_executive != user and not _is_finance_head(user):
                raise PermissionDenied(f"This payout is assigned to {obj.final_executive.name}.")

            if data:
                payment_mode = data.get('payment_mode')
                transaction_id = data.get('transaction_id')
                amount = getattr(obj, 'executive_approved_amount', getattr(obj, 'approved_amount', 0))
                
                # Capture details
                if hasattr(obj, 'payment_mode'): obj.payment_mode = payment_mode
                if hasattr(obj, 'transaction_id'): obj.transaction_id = transaction_id
                if hasattr(obj, 'payment_date'): obj.payment_date = data.get('payment_date') or timezone.now()
                if hasattr(obj, 'finance_remarks'): obj.finance_remarks = data.get('remarks', '')
                if hasattr(obj, 'processed_by'): obj.processed_by = user
            
            # Set target status based on model type
            if isinstance(obj, TravelClaim):
                obj.status = 'Paid'
            else:
                obj.status = 'COMPLETED'
                
            obj.current_approver = None
            obj.save()
            
            # Record Lifecycle Event
            trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
            if trip:
                # If this is a claim settlement, also update the trip status to 'Settled'
                if isinstance(obj, TravelClaim):
                    trip.status = 'Settled'
                    trip.save()
                update_trip_lifecycle(trip, "Settlement", f"Final Transfer completed by {user.name}.")
            
            Notification.objects.create(
                user=requester,
                title=f"Amount Credited [{obj.trip_id}]",
                message=f"Your {request_type} {obj.trip_id} has been fully approved and the amount has been credited to your account.",
                type='success'
            )
            return Response({"message": f"{action} completed and phase closed."})

        if action == 'RejectByFinance':
            obj.status = 'Rejected by Finance'
            reason = data.get('remarks', 'No reason provided') if data else ""
            if hasattr(obj, 'finance_remarks'): obj.finance_remarks = reason
            obj.save()
            Notification.objects.create(
                user=requester,
                title=f"Finance: Request Rejected [{obj.trip_id}]",
                message=f"Your {request_type} {obj.trip_id} was rejected by Finance. Reason: {reason}",
                type='error'
            )
            return Response({"message": "Rejected by Finance"})

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsCustomAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put', 'head', 'options']

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Expense.objects.none()
            
        role_name = user.role.name.lower() if hasattr(user, 'role') else ''
        is_admin = role_name in ['admin', 'superuser']
        is_finance = 'finance' in role_name or role_name == 'cfo'
        is_manager = role_name == 'reporting_authority'
        
        if is_admin or is_finance:
            queryset = self.queryset
        elif is_manager:
            # Allow managers to see their own expenses OR expenses for trips they are/were responsible for
            # We use snapshot names for efficient filtering as dynamic hierarchy lookup is too slow for list views
            queryset = self.queryset.filter(
                Q(trip__user=user) | 
                Q(trip__current_approver=user) |
                Q(trip__reporting_manager_name=user.name) |
                Q(trip__senior_manager_name=user.name) |
                Q(trip__hod_director_name=user.name)
            )
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        serializer.save()

class PolicyDocumentViewSet(viewsets.ModelViewSet):
    queryset = PolicyDocument.objects.all()
    serializer_class = PolicyDocumentSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PolicyDocumentDetailSerializer
        return PolicyDocumentSerializer

    def get_queryset(self):
        # Always return a fresh queryset; returning the class-level queryset
        # directly can reuse a cached empty result across requests.
        return PolicyDocument.objects.all()

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(uploaded_by=user)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from core.permissions import IsAdmin
            return [IsAdmin()]
        return [IsCustomAuthenticated()]

class TravelClaimViewSet(viewsets.ModelViewSet):
    queryset = TravelClaim.objects.all()
    serializer_class = TravelClaimSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelClaim.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        
        reporting_manager = user.reporting_manager
        senior_manager = user.senior_manager
        hod_director = user.hod_director
        
        # Logic to find first available approver
        current_approver = reporting_manager
        h_level = 1
        
        if not current_approver:
            current_approver = senior_manager
            h_level = 2
        
        if not current_approver:
            current_approver = hod_director
            h_level = 3
            
        if not current_approver:
            # If no managers at all, go to HR
            current_approver = get_hr_head(user)

        from django.db.models import Sum
        total_expense_sum = trip.expenses.aggregate(s=Sum('amount'))['s'] or 0

        claim = serializer.save(
            status='Submitted',
            current_approver=current_approver,
            hierarchy_level=h_level,
            submitted_at=timezone.now(),
            total_amount=total_expense_sum,
            approved_amount=total_expense_sum,
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager=reporting_manager,
            senior_manager=senior_manager,
            hod_director=hod_director,
            reporting_manager_name=reporting_manager.name if reporting_manager else None,
            senior_manager_name=senior_manager.name if senior_manager else None,
            hod_director_name=hod_director.name if hod_director else None
        )
        
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title=f"New Expense Claim [{claim.trip.trip_id}]",
                message=f"{user.name} has submitted an expense claim for Trip {claim.trip.trip_id}.",
                type='info'
            )
            
            # NOTIFY USER: Claim submitted
            Notification.objects.create(
                user=user,
                title=f"Expense Claim Sent [{claim.trip.trip_id}]",
                message=f"Your expense claim for Trip {claim.trip.trip_id} has been submitted and sent for approval.",
                type='success'
            )
            
        # Notify HR
        notify_hr("New Expense Claim", f"{user.name} has submitted an expense claim for Trip {claim.trip.trip_id}.")
            
        self._update_trip_lifecycle(claim.trip)

    def perform_update(self, serializer):
        claim = serializer.save()
        
        from django.db.models import Sum
        total_amount = claim.trip.expenses.aggregate(s=Sum('amount'))['s'] or 0
        claim.total_amount = total_amount
        claim.approved_amount = total_amount
        claim.save()
        
        if claim.status == 'Submitted':
            self._update_trip_lifecycle(claim.trip)

    def _update_trip_lifecycle(self, trip):
        update_trip_lifecycle(trip, "Settlement", "Travel reimbursement claim submitted for review.")

class TravelAdvanceViewSet(viewsets.ModelViewSet):
    queryset = TravelAdvance.objects.all()
    serializer_class = TravelAdvanceSerializer

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelAdvance.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
            
        reporting_manager = user.reporting_manager
        senior_manager = user.senior_manager
        hod_director = user.hod_director

        # Logic to find first available approver
        current_approver = reporting_manager
        h_level = 1
        
        if not current_approver:
            current_approver = senior_manager
            h_level = 2
        
        if not current_approver:
            current_approver = hod_director
            h_level = 3
            
        if not current_approver:
            # If no managers at all, go to HR
            current_approver = get_hr_head(user)

        advance = serializer.save(
            status='Submitted',
            current_approver=current_approver,
            hierarchy_level=h_level,
            submitted_at=timezone.now(),
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager=reporting_manager,
            senior_manager=senior_manager,
            hod_director=hod_director,
            reporting_manager_name=reporting_manager.name if reporting_manager else None,
            senior_manager_name=senior_manager.name if senior_manager else None,
            hod_director_name=hod_director.name if hod_director else None
        )
        
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title=f"New Advance Request [{advance.trip.trip_id}]",
                message=f"{user.name} has requested an advance for Trip {advance.trip.trip_id}.",
                type='info'
            )
            
            # NOTIFY USER: Advance submitted
            Notification.objects.create(
                user=user,
                title=f"Advance Request Sent [{advance.trip.trip_id}]",
                message=f"Your advance request for Trip {advance.trip.trip_id} has been submitted and sent for approval.",
                type='success'
            )
            
        # Notify HR
        notify_hr("New Advance Request", f"{user.name} has requested an advance of ₹{advance.requested_amount} for Trip {advance.trip.trip_id}.")

        self._update_trip_lifecycle(advance)

    def perform_update(self, serializer):
        advance = serializer.save()
        if advance.status == 'Submitted':
            self._update_trip_lifecycle(advance)

    def _update_trip_lifecycle(self, advance):
        update_trip_lifecycle(advance.trip, "Advance Requested", f"Pre-trip advance of ₹{advance.requested_amount} requested for: {advance.purpose[:40]}...")



class DisputeViewSet(viewsets.ModelViewSet):
    queryset = Dispute.objects.all()
    serializer_class = DisputeSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Dispute.objects.none()
        
        if hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']:
            return Dispute.objects.all()
        
        return Dispute.objects.filter(raised_by=user)

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(raised_by=user)

class TripOdometerViewSet(viewsets.ModelViewSet):
    queryset = TripOdometer.objects.all()
    serializer_class = TripOdometerSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TripOdometer.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
            
            
        odo = serializer.save()
        if odo.start_odo_reading:
            trip = odo.trip
            trip.status = 'On-Going'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Started", f"Trip journey started with odometer reading {odo.start_odo_reading} KM.")

    def perform_update(self, serializer):
        odo = serializer.save()
        if odo.end_odo_reading:
            trip = odo.trip
            trip.status = 'Completed'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Ended", f"Trip journey completed with final odometer reading {odo.end_odo_reading} KM.")

class DashboardStatsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        is_admin = user.role.name.lower() == 'admin'
        is_gh_manager = user.role.name.lower() == 'guesthousemanager'
        is_fin_head = _is_finance_head(user)
        is_fin_exec = _is_finance_executive(user)
        is_finance = is_fin_head or is_fin_exec
        is_cfo = 'cfo' in (user.role.name.lower() if user.role else '')
        is_hr = _is_hr(user)

        if is_admin or is_gh_manager or is_finance or is_cfo:
            trips = Trip.objects.all()
            base_expenses = Expense.objects.all()
        else:
            trips = Trip.objects.filter(user=user)
            base_expenses = Expense.objects.filter(trip__user=user)

        total_trips = trips.count()
        in_review = trips.filter(status='Pending').count()
        
        # Count tasks awaiting this user's approval
        pending_action = 0
        # variables already defined above

        if not is_admin:
            pending_action += Trip.objects.filter(current_approver=user).count()
            pending_action += TravelAdvance.objects.filter(current_approver=user).count()
            pending_action += TravelClaim.objects.filter(current_approver=user).count()
            
            # For Finance, count specific stages they act on
            if is_fin_head:
                pending_action += TravelAdvance.objects.filter(status='PENDING_HEAD').count()
                pending_action += TravelClaim.objects.filter(status='PENDING_HEAD').count()
            elif is_fin_exec:
                money_statuses = ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE']
                pending_action += TravelAdvance.objects.filter(status__in=money_statuses).count()
                pending_action += TravelClaim.objects.filter(status__in=money_statuses).count()
            elif is_hr:
                pending_action += Trip.objects.filter(status='Manager Approved').count()
            elif is_gh_manager:
                # For GH Manager, count trips that need room booking
                pending_action += Trip.objects.filter(
                    accommodation_requests__contains='Request for Room',
                    status__in=['Manager Approved', 'Approved']
                ).exclude(room_bookings__isnull=False).distinct().count()
        
        total_expenses = base_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        
        # compute total approved advances for the selected trips - used for wallet/advance balances
        # mirror TripSerializer.get_total_approved_advance logic (consider executive_approved_amount when >0)
        total_approved_advances = 0.0
        advances_qs = TravelAdvance.objects.filter(
            trip__in=trips,
            status__in=['Paid', 'Transferred', 'COMPLETED']
        )
        for adv in advances_qs:
            amt = float(adv.executive_approved_amount) if float(adv.executive_approved_amount) > 0 else float(adv.requested_amount)
            total_approved_advances += amt
        
        wallet_balance = float(total_approved_advances) - float(total_expenses)
        
        approved_expenses_qs = base_expenses.filter(
            Q(trip__claim__status__in=['Approved', 'Paid']) |
            Q(trip__status__in=['Approved', 'Completed', 'Settled'], trip__consider_as_local=True)
        )
        approved_expenses = approved_expenses_qs.aggregate(Sum('amount'))['amount__sum'] or 0
        
        categories = base_expenses.values('category').annotate(total=Sum('amount'))
        
        recent_trips = trips.order_by('-created_at')[:5]
        recent_data = []
        for t in recent_trips:
            # If trip is approved/completed, show actual total instead of estimate
            if t.status in ['Approved', 'Completed', 'Settled']:
                actual_total = Expense.objects.filter(trip=t).aggregate(Sum('amount'))['amount__sum'] or 0
                display_amount = f"₹{actual_total:,.0f}" if actual_total > 0 else t.cost_estimate
            else:
                display_amount = t.cost_estimate

            recent_data.append({
                "id": t.trip_id,
                "title": f"{'Travel' if t.consider_as_local else 'Trip'} to {t.destination}",
                "subtitle": f"{t.user.name} - {t.purpose}" if (is_admin or is_gh_manager) and t.user else t.purpose,
                "status": t.status,
                "amount": display_amount
            })

        kpis = [
            { "title": 'Total Trips', "value": str(total_trips), "label": 'Managed trips' if is_admin or is_finance or is_gh_manager else 'Your trips', "icon": 'Briefcase', "color": 'orange' },
            { "title": 'Approved Expenses', "value": f"₹{approved_expenses:,.0f}", "label": 'Finalized' if is_finance else 'Confirmed', "icon": 'CreditCard', "color": 'red' },
            { "title": 'Total Spend' if not is_finance else 'Total Disbursements', "value": f"₹{total_expenses:,.0f}", "label": 'Recorded', "icon": 'TrendingUp', "color": 'magenta' },
            { 
                "title": 'Action Required' if pending_action > 0 else 'In Review', 
                "value": str(pending_action if pending_action > 0 else in_review), 
                "label": 'Pending your action' if pending_action > 0 else 'Pending trips', 
                "icon": 'Clock', 
                "color": 'yellow' 
            }
        ]

        if is_finance:
            # Shift to vibrant functional colors (strictly within supported index.css classes)
            kpis[0]['color'] = 'orange'
            kpis[1]['color'] = 'magenta'
            kpis[1]['title'] = 'Net Payouts'
            kpis[2]['color'] = 'red'

        return Response({
            "kpis": kpis,
            "recent_activity": recent_data,
            "is_finance_hub": is_finance,
            "expenditure_mix": [
                { 
                    "type": dict(Expense.CATEGORY_CHOICES).get(cat['category'], cat['category']), 
                    "amount": float(cat['total']), 
                    "percentage": (float(cat['total']) / float(total_expenses) * 100) if total_expenses > 0 else 0 
                }
                for cat in categories
            ],
            "total_spend": total_expenses,
            # wallet_balance represents approved advances minus expenses across the user/trips
            "wallet_balance": wallet_balance,
            # advance_balance is simply the sum of approved advances
            "advance_balance": float(total_approved_advances)
        })

class TripSettlementView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        trip_id_enc = request.query_params.get('trip_id')
        
        if not trip_id_enc:
            # If no trip_id provided, return all trips that are ready for settlement
            # (i.e. trips with approved claims that aren't PAID/COMPLETED yet)
            # For simplicity, returning a list of all trips the user can see for now
            # but we filter for finance if they are finance
            is_finance = _is_finance_executive(user) or _is_finance_head(user)
            if is_finance:
                trips = Trip.objects.filter(status__in=['Claim Submitted', 'Manager Approved', 'Approved'])
            else:
                trips = Trip.objects.filter(user=user)
            
            data = []
            for t in trips:
                advances = TravelAdvance.objects.filter(trip=t, status='COMPLETED').aggregate(Sum('requested_amount'))['requested_amount__sum'] or 0
                claim_amt = 0
                if hasattr(t, 'claim'):
                    claim_amt = t.claim.total_amount
                
                data.append({
                    "trip_id": t.trip_id,
                    "destination": t.destination,
                    "employee": t.user.name if t.user else "Unknown",
                    "advance": float(advances),
                    "claim": float(claim_amt),
                    "balance": float(claim_amt - advances),
                    "status": t.status
                })
            return Response(data)

        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        advances_list = TravelAdvance.objects.filter(trip=trip).order_at = ['created_at']
        total_advance = sum(a.requested_amount for a in advances_list if a.status == 'COMPLETED')
        
        claim_amt = 0
        claim_id = None
        claim_date = None
        if hasattr(trip, 'claim'):
            claim_amt = trip.claim.total_amount
            claim_id = f"CLAIM-{trip.claim.id}"
            claim_date = trip.claim.submitted_at.strftime("%B %d, %Y") if trip.claim.submitted_at else None

        breakdown = []
        for adv in advances_list:
            if adv.status == 'COMPLETED':
                breakdown.append({
                    "id": f"ADV-{adv.id}",
                    "type": "Advance",
                    "description": f"Advance via {adv.payment_mode or 'Bank Transfer'}",
                    "date": adv.payment_date.strftime("%B %d, %Y") if adv.payment_date else adv.created_at.strftime("%B %d, %Y"),
                    "amount": -float(adv.requested_amount),
                    "is_negative": True
                })

        if claim_amt > 0:
            breakdown.append({
                "id": claim_id,
                "type": "Claim",
                "description": f"Expense Claim: {trip.trip_id}",
                "date": claim_date or trip.updated_at.strftime("%B %d, %Y"),
                "amount": float(claim_amt),
                "is_negative": False
            })

        balance = float(claim_amt - total_advance)
        
        return Response({
            "summary": {
                "advance": float(total_advance),
                "claimTotal": float(claim_amt),
                "balance": balance,
                "type": "Payable" if balance > 0 else "Recoverable",
                "status": trip.status
            },
            "breakdown": breakdown,
            "trip": {
                "id": trip.trip_id,
                "destination": trip.destination,
                "employee": trip.user.name if trip.user else "Unknown"
            }
        })

    def post(self, request):
        # Handle Finalize & Settle action
        user = getattr(request, 'custom_user', None)
        trip_id_enc = request.data.get('trip_id')
        if not trip_id_enc:
            return Response({"error": "Trip ID required"}, status=400)
            
        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        # Logic to finalize
        trip.status = 'Settled'
        trip.save()
        
        if hasattr(trip, 'claim'):
            trip.claim.status = 'Paid'
            trip.claim.save()

        update_trip_lifecycle(trip, "Final Settlement", f"Accounts settled and closed by {user.name}.")
        
        return Response({"message": "Trip account successfully settled."})

class CFOWarRoomView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        now = timezone.now()
        first_day_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # 1. Monthly Total Spend
        monthly_spend = Expense.objects.filter(date__gte=first_day_of_current_month.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Comparison with last month
        last_month_end = first_day_of_current_month - timezone.timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_spend = Expense.objects.filter(date__gte=last_month_start.date(), date__lte=last_month_end.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        percent_change_spend = 0
        if last_month_spend > 0:
            percent_change_spend = ((float(monthly_spend) - float(last_month_spend)) / float(last_month_spend)) * 100
        
        # 2. Avg Cost Per Trip
        total_claims = TravelClaim.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_trips = Trip.objects.count()
        avg_cost = (float(total_claims) / total_trips) if total_trips > 0 else 0
        
        # 3. Guest House Occupancy
        from guest_house.models import Room, RoomBooking
        total_rooms = Room.objects.count()
        # Active bookings today
        active_bookings = RoomBooking.objects.filter(start_date__lte=now, end_date__gte=now).count()
        occupancy = (active_bookings / total_rooms * 100) if total_rooms > 0 else 0
        
        # 4. Policy Overrides/Disputes
        # Using Category 'Policy' in Dispute as a proxy for overrides
        policy_issues = Dispute.objects.filter(category='Policy').count()
        policy_rate = (policy_issues / total_trips * 100) if total_trips > 0 else 0

        # KPI Stats Construction
        stats = [
            { 
                "title": 'Total Spend (Monthly)', 
                "value": f"₹{monthly_spend:,.0f}", 
                "change": f"{abs(percent_change_spend):.1f}%", 
                "trend": 'up' if percent_change_spend >= 0 else 'down',
                "icon": 'IndianRupee'
            },
            { 
                "title": 'Avg Cost per Trip', 
                "value": f"₹{avg_cost:,.0f}", 
                "change": '+2.4%', # Placeholder for complexity
                "trend": 'up', 
                "icon": 'TrendingUp' 
            },
            { 
                "title": 'Guest House Occupancy', 
                "value": f"{occupancy:.1f}%", 
                "change": '+1.8%', 
                "trend": 'up', 
                "icon": 'Building2' 
            },
            { 
                "title": 'Policy Overrides', 
                "value": f"{policy_rate:.1f}%", 
                "change": '-0.2%', 
                "trend": 'down', 
                "icon": 'AlertCircle' 
            },
        ]

        # 5. Spend by Department
        dept_raw = Expense.objects.values('trip__user_department').annotate(total=Sum('amount')).order_by('-total')
        spend_by_dept = [
            { "name": d['trip__user_department'] or "Unknown", "value": float(d['total']) }
            for d in dept_raw
        ]

        # 6. Advance Aging
        advances = TravelAdvance.objects.exclude(status='COMPLETED')
        aging_buckets = { "0-30": 0, "31-60": 0, "60+": 0 }
        for adv in advances:
            days = (now - adv.created_at).days
            amt = float(adv.requested_amount)
            if days <= 30: aging_buckets["0-30"] += amt
            elif days <= 60: aging_buckets["31-60"] += amt
            else: aging_buckets["60+"] += amt
            
        aging_data = [
            { "range": '0-30 Days', "amount": aging_buckets["0-30"], "color": 'success' },
            { "range": '31-60 Days', "amount": aging_buckets["31-60"], "color": 'warning' },
            { "range": '60+ Days', "amount": aging_buckets["60+"], "color": 'danger' },
        ]

        # 7. Critical Anomalies
        anomalies = [
            { "entity": "Logistics Surge", "reason": "Local Conveyance in Metro areas up by 40%", "impact": "High", "action": "Review" },
            { "entity": "Dept Override", "reason": "Bulk travel booking policy bypassed in IT dept", "impact": "Medium", "action": "Investigate" }
        ]

        return Response({
            "stats": stats,
            "report_month": now.strftime("%B %Y"),
            "spend_by_dept": spend_by_dept,
            "aging": aging_data,
            "anomalies": anomalies
        })

import json

class BulkActivityBatchViewSet(viewsets.ModelViewSet):
    queryset = BulkActivityBatch.objects.all()
    serializer_class = BulkActivityBatchSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user: return BulkActivityBatch.objects.none()
        
        role_name = (user.role.name if user.role else '').lower()
        
        # Admins/Finance/COO see all
        if any(kw in role_name for kw in ['admin', 'finance', 'cfo', 'coo']):
            return self.queryset
        
        # Allow any user to see batches they submitted OR batches they need to approve.
        # This ensures that managers (like a COO) can see and approve their team's uploads 
        # even if their system role is simply 'Employee'.
        return self.queryset.filter(Q(user=user) | Q(current_approver=user)).order_by('-created_at')

    @action(detail=False, methods=['get'])
    def template(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import FormulaRule
        import datetime
        from django.utils import timezone

        wb = openpyxl.Workbook()


        # ── Fetch Locations (Inclusive: Self + Team + Project) ──
        user = getattr(request, 'custom_user', None)
        manager_code = user.employee_id if user else None
        project_code = user.project_code if user else 'N/A'
        
        locations_set = set()
        
        # 1. Own Location
        if user and user.office_location:
            locations_set.add(user.office_location.strip())
            
        # 2. Team Locations ("Child Users") - Recursive fetch via API
        if manager_code:
            from api_management.services import get_manager_reports_locations
            try:
                # This fetches unique clusters/districts for the entire reporting chain
                team_locs = get_manager_reports_locations(manager_code)
                for loc in team_locs:
                    if loc: locations_set.add(loc.strip())
            except Exception as e:
                print(f"Template Location Sync Error: {e}")

        # 3. Project Jurisdiction Locations
        if project_code and project_code != 'N/A':
            from travel_masters.models import Location
            juris_locs = Location.objects.filter(
                jurisdiction_districts__project_code=project_code
            ).values_list('name', flat=True).distinct()
            for loc in juris_locs:
                if loc: locations_set.add(loc.strip())
        
        # Fallback to all Districts if the set is empty or very small
        if len(locations_set) < 2:
            from travel_masters.models import Location
            all_distruits = Location.objects.filter(location_type__iexact='District').values_list('name', flat=True).order_by('name')[:100]
            for loc in all_distruits:
                if loc: locations_set.add(loc.strip())

        # Final ordered list
        locations = sorted(list(locations_set)) if locations_set else ["Head Office", "Field Office", "Client Site"]

        # locations list is ready; will be written to Column Z after ws is created

        # ── Main data sheet ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Monthly Activities"

        # ── Styles ──────────────────────────────────────────────────────────
        HEADER_FILL   = PatternFill("solid", fgColor="FFBB0633")   # brand red (8-char)
        NOTE_FILL     = PatternFill("solid", fgColor="FFFFF3CD")   # warn yellow
        HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
        NOTE_FONT     = Font(name="Calibri", italic=True, color="856404", size=9)
        DATA_FONT     = Font(name="Calibri", size=10)
        CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin          = Side(style="thin", color="CCCCCC")
        BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Column layout ────────────────────────────────────────────────────
        # A=Date  B=Time  C=From Location  D=To Location  E=Purpose
        columns = [
            ("Date",          16, CENTER),
            ("Time",          14, CENTER),
            ("From Location", 28, LEFT),
            ("To Location",   28, LEFT),
            ("Purpose",       40, LEFT),
        ]

        # Row 1 – Headers
        ws.row_dimensions[1].height = 30
        for col_idx, (label, width, align) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row 2 – Instructional note (merged across all columns)
        ws.merge_cells("A2:E2")
        note_cell = ws.cell(row=2, column=1,
            value="📋  Instructions: Date must be ≥ today  |  Time must be between 06:00 and 23:59 (24h)  |  "
                  "Choose From/To Location from the dropdown  |  Purpose is free text")
        note_cell.font      = NOTE_FONT
        note_cell.fill      = NOTE_FILL
        note_cell.alignment = LEFT
        ws.row_dimensions[2].height = 22

        # Row 3 – Sample data row
        today_str = datetime.date.today().strftime("%Y-%m-%d")

        # Explicitly convert to IST (UTC+5:30) — server TIME_ZONE is UTC
        from zoneinfo import ZoneInfo
        ist_tz    = ZoneInfo('Asia/Kolkata')
        now_ist   = datetime.datetime.now(ist_tz)
        # Use 09:30 as a standard sample time rather than dynamic 'now' to avoid confusing users
        sample_time = "09:30"

        # ── Sample Data (Row 3) ──
        # Use real objects for comparison safety in Excel formulas
        sample = [
            datetime.date.today(), 
            datetime.time(9, 30), 
            locations[0] if locations else "Head Office",
            locations[1] if len(locations) > 1 else "Field Office",
            "Site Inspection / Field Visit"
        ]
        ws.row_dimensions[3].height = 20
        for col_idx, val in enumerate(sample, start=1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font      = DATA_FONT
            cell.alignment = columns[col_idx - 1][2]
            cell.border    = BORDER
            if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
            if col_idx == 2: cell.number_format = 'HH:MM'

        # Rows 4-503 – Empty data rows (500 rows)
        for row in range(4, 504):
            ws.row_dimensions[row].height = 20
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.font      = DATA_FONT
                cell.alignment = columns[col_idx - 1][2]
                cell.border    = BORDER
                if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
                if col_idx == 2: cell.number_format = 'HH:MM'

        DATA_ROWS = "3:503"   # applies to sample + 500 blank rows

        # ── Location source list in Column Z (same-sheet = no corruption ever) ──
        # Column Z is set very narrow so it is effectively invisible to users.
        ws.column_dimensions['Z'].width = 0.1
        for i, loc in enumerate(locations, start=1):
            ws.cell(row=i, column=26, value=str(loc))

        # Create a Named Range pointing to Column Z on THIS sheet.
        # Named ranges scroll correctly in Excel dropdowns even with 100+ items.
        from openpyxl.workbook.defined_name import DefinedName
        sheet_title = ws.title
        loc_named_range = DefinedName(
            'LOC_LIST',
            attr_text=f"'{sheet_title}'!$Z$1:$Z${len(locations)}"
        )
        wb.defined_names.add(loc_named_range)

        # ── Validation 1: Date >= today ──────────────────────────────────────
        today_serial = (datetime.date.today() - datetime.date(1899, 12, 30)).days
        dv_date = DataValidation(
            type="date", 
            operator="greaterThanOrEqual", 
            formula1=str(today_serial),
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Date",
            error="The date cannot be earlier than today."
        )
        ws.add_data_validation(dv_date)
        dv_date.add("A3:A503")

        # ── Validation 2a: Time for Row 3 (first data row – no previous row) ──
        dv_time_first = DataValidation(
            type="custom",
            formula1="=AND(B3>=0.25,B3<=0.999)",
            showErrorMessage=True, errorStyle="stop",
            errorTitle="Invalid Time",
            error="Time must be between 06:00 and 23:59 (e.g. 09:30)."
        )
        ws.add_data_validation(dv_time_first)
        dv_time_first.add("B3")

        # ── Validation 2b: Time for Rows 4–503 ────────────────────────────────
        # Rules: in range 06:00–23:59 AND at least 10 minutes after previous on same day.
        # 10 minutes = 10/1440 in Excel decimal time.
        dv_time_seq = DataValidation(
            type="custom",
            formula1="=AND(B4>=0.25,B4<=0.999,OR(A4>A3,B4>B3+(10/1440)))",
            showErrorMessage=True, errorStyle="stop",
            errorTitle="Time Sequence Error",
            error="Time must be at least 10 minutes after the previous row on the same day, or start a new date."
        )
        ws.add_data_validation(dv_time_seq)
        dv_time_seq.add("B4:B503")

        # ── Validation 3a: From Location – dropdown list ───────────────────────
        dv_from = DataValidation(
            type="list", formula1="LOC_LIST", showDropDown=False,
            showErrorMessage=True, errorStyle="stop",
            errorTitle="Invalid From Location",
            error="Please select a From location from the dropdown."
        )
        ws.add_data_validation(dv_from)
        dv_from.add("C3:C503")

        # ── Validation 3b: To Location – custom STOP (in-list AND != From) ───
        # ── Validation 3b: To Location – dropdown list (dropdown arrow required) ─
        # NOTE: Excel only allows ONE validation per cell. To show the dropdown arrow,
        # we MUST use type="list". Same-From/To is handled via conditional formatting below.
        dv_to = DataValidation(
            type="list", formula1="LOC_LIST", showDropDown=False,
            showErrorMessage=True, errorStyle="stop",
            errorTitle="Invalid To Location",
            error="Please select a To location from the dropdown."
        )
        ws.add_data_validation(dv_to)
        dv_to.add("D3:D503")

        # ── Conditional Formatting: visual alerts ─────────────────────────────
        # ORANGE: From == To on same row (same location trip is invalid)
        orange_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
        ws.conditional_formatting.add("D3:D503",
            FormulaRule(formula=["AND(D3<>\"\",D3=C3)"], fill=orange_fill))
        # RED: route continuity break (From ≠ prev row's To on same day)
        red_fill = PatternFill(start_color="FFE99696", end_color="FFE99696", fill_type="solid")
        ws.conditional_formatting.add("C4:C503",
            FormulaRule(formula=["AND(A4=A3,C4<>D3,C4<>\"\")"], fill=red_fill))

        # ── Validation 4: Purpose ────────────────────────────────────────────
        dv_purpose = DataValidation(
            type="textLength", 
            operator="greaterThan", 
            formula1="0",
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Missing Info",
            error="Purpose of visit is mandatory."
        )
        ws.add_data_validation(dv_purpose)
        dv_purpose.add("E3:E503")

        # ── Freeze panes below header + note rows ────────────────────────────
        ws.freeze_panes = "A3"

        # ── Output ───────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bulk_activity_template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def upload(self, request):
        user = getattr(request, 'custom_user', None)
        file = request.FILES.get('file')
        trip_id = request.data.get('trip_id') # Selected by user in UI
        
        if not file:
            return Response({"error": "No file uploaded"}, status=400)
            
        try:
            import pandas as pd
            # Explicitly use openpyxl but with a generic fallback
            try:
                xls = pd.ExcelFile(file)
            except Exception:
                try:
                    xls = pd.ExcelFile(file, engine='openpyxl')
                except Exception as ex:
                    return Response({"error": f"Supported Excel engines failed: {str(ex)}"}, status=400)
            
            df_full = None
            sheet_name = None
            found_header = False
            header_idx = 0
            
            # Try sheets in order: look for the one containing "Date" or "Activity"
            for s in xls.sheet_names:
                try:
                    temp_df = xls.parse(s, header=None)
                except Exception:
                    continue
                if temp_df.empty: continue
                
                # Scan first 20 rows for "Date"
                for idx, row in temp_df.head(20).iterrows():
                    row_vals = [str(c).strip().lower() for c in row if pd.notna(c)]
                    if 'date' in row_vals or 'activity' in row_vals or 'plan' in row_vals:
                        df_full = temp_df
                        header_idx = idx
                        found_header = True
                        sheet_name = s
                        break
                if found_header: break
            
            if not found_header or df_full is None:
                # Secondary fallback: try first sheet anyway 
                sheet_name = xls.sheet_names[0]
                df_full = xls.parse(sheet_name, header=None)
                header_idx = 0 
            
            # 2. Slice and Map Columns
            header_row_raw = df_full.iloc[header_idx]
            df = df_full.iloc[header_idx:].copy()
            
            # Handle duplicate or empty headers
            new_cols = []
            for i, c in enumerate(header_row_raw):
                c_str = str(c).strip() if pd.notna(c) else ""
                if not c_str:
                    new_cols.append(f"Col_{i}")
                else:
                    new_cols.append(c_str)
            df.columns = new_cols
            df = df.iloc[1:] # Data starts after header row
            
            # Improved Case-insensitive column matcher
            col_map = {}
            for col in df.columns:
                c_str = str(col).strip().lower()
                if 'date' in c_str: col_map['date'] = col
                elif 'time' in c_str: col_map['time'] = col
                elif 'from' in c_str or 'origin' in c_str or 'source' in c_str or 'src' in c_str: col_map['from'] = col
                elif 'to' in c_str or 'dest' in c_str: col_map['to'] = col
                elif 'purp' in c_str or 'intent' in c_str: col_map['purpose'] = col

            # CRITICAL: Validate that all required columns are MAPPED
            required_keys = ['date', 'time', 'from', 'to', 'purpose']
            missing_cols = [k for k in required_keys if k not in col_map]
            if missing_cols:
                return Response({
                    "error": f"Missing mandatory columns in Excel: {', '.join(missing_cols)}. Please check headers.",
                    "cols_found": list(df.columns)
                }, status=400)

            rows = []
            last_date = None
            last_to = None
            last_time = None
            last_excel_row = None

            for i, row in df.iterrows():
                # Skip absolute blanks
                if row.dropna().empty:
                    continue
                
                # Primary data extraction
                curr_date = str(row.get(col_map.get('date'))).strip()
                if len(curr_date) > 10: curr_date = curr_date[:10]
                
                # Skip header/instruction repetitions (CRITICAL: Do this BEFORE validation)
                if any(kw in curr_date.lower() for kw in ['instruc', 'sample', 'yyyy', 'month', 'date']):
                    continue

                # If DATE field is blank, treat as end-of-data and stop parsing
                date_val = row.get(col_map.get('date'))
                if pd.isna(date_val) or str(date_val).strip() in ('', 'nan', 'NaT', 'None'):
                    # If we already collected rows, stop cleanly.
                    # If no rows yet, just skip (might be a gap before data starts).
                    if rows:
                        break
                    continue

                # Check for mandatory data in this row (skip partial rows silently)
                missing_data = [
                    k for k in required_keys
                    if pd.isna(row.get(col_map.get(k))) or str(row.get(col_map.get(k))).strip() in ('', 'nan', 'NaT', 'None')
                ]
                if missing_data:
                    # Skip rows that have a date but are missing other fields (gaps in middle)
                    continue


                curr_from = str(row.get(col_map.get('from'))).strip()
                curr_to = str(row.get(col_map.get('to'))).strip()
                
                # Extract and normalize time (handle 08.30, 8:30, 08;30)
                time_raw = row.get(col_map.get('time'))
                time_str = str(time_raw).replace('.', ':').replace(';', ':').replace(' ', '').strip()
                
                # Try handling Excel decimal serials (fraction of day)
                try:
                    frac = float(time_str)
                    if 0 <= frac < 1:
                        total_secs = int(frac * 86400)
                        time_str = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                except:
                    pass
                
                # Ensure HH:MM padding for string comparison robustness (08:30 vs 09:30)
                if ':' in time_str:
                    parts = time_str.split(':')
                    try:
                        h = int(parts[0])
                        m = int(parts[1])
                        time_str = f"{h:02d}:{m:02d}"
                    except:
                        pass
                else: 
                    # If it's just "8", make it "08:00"
                    if time_str.isdigit():
                        time_str = f"{int(time_str):02d}:00"

                # ── LOGICAL MULTI-ROW VALIDATION ──
                excel_row_ref = i + 1
                if last_date == curr_date:
                    # 1. Route Continuity Check
                    if last_to and curr_from and last_to.lower() != curr_from.lower():
                        return Response({
                            "error": f"Logical Error on {curr_date}: Route mismatch between Row {last_excel_row} and Row {excel_row_ref}. "
                                     f"Row {excel_row_ref} must start from '{last_to}' (the previous destination), but starts from '{curr_from}'."
                        }, status=400)
                    
                    # 2. Time Sequence Check
                    if last_time and time_str < last_time:
                         return Response({
                            "error": f"Logical Error on {curr_date}: Time sequence mismatch between Row {last_excel_row} and Row {excel_row_ref}. "
                                     f"Row {excel_row_ref} time ({time_str}) cannot be earlier than Row {last_excel_row} time ({last_time})."
                        }, status=400)

                # Update trackers for next row
                last_date = curr_date
                last_to = curr_to
                last_time = time_str
                last_excel_row = excel_row_ref

                # Append record
                rows.append({
                    "date": curr_date,
                    "time": time_str,
                    "mode": "Bike",
                    "origin_route": curr_from,
                    "destination_route": curr_to,
                    "vehicle": "Own Bike",
                    "visit_intent": str(row.get(col_map.get('purpose'))).strip()
                })

            if not rows:
                debug_info = {
                    "sheet": sheet_name,
                    "header_row_idx": header_idx,
                    "cols_found": list(df.columns),
                    "mapped": col_map
                }
                return Response({
                    "error": f"No valid data rows found in sheet '{sheet_name}'. Please ensure data is filled starting from the row after the headers.",
                    "debug_info": debug_info
                }, status=400)
            
            # Management Hierarchy Snapshots
            rm = user.reporting_manager
            sm = user.senior_manager
            hod = user.hod_director
            
            def is_management_role(u):
                if not u or not u.role: return False
                return u.role.name.lower() in ['admin', 'it-admin', 'superuser', 'it admin', 'system administrator']

            current_approver = rm if not is_management_role(rm) else None
            h_level = 1
            
            if not current_approver:
                current_approver = sm if not is_management_role(sm) else None
                h_level = 2
            
            if not current_approver:
                current_approver = hod if not is_management_role(hod) else None
                h_level = 3
                
            if not current_approver:
                 from .views import get_hr_head
                 current_approver = get_hr_head(user)
                 h_level = 1

            batch = BulkActivityBatch.objects.create(
                user=user,
                trip_id=trip_id,
                file_name=getattr(file, 'name', 'Uploaded_Log.xlsx') or 'Uploaded_Log.xlsx',
                data_json=rows,
                status='Submitted',
                current_approver=current_approver,
                hierarchy_level=h_level,
                reporting_manager=rm,
                senior_manager=sm,
                hod_director=hod,
                reporting_manager_name=rm.name if rm else None,
                senior_manager_name=sm.name if sm else None,
                hod_director_name=hod.name if hod else None
            )
            
            if current_approver:
                Notification.objects.create(
                    user=current_approver,
                    title=f"New Bulk Activity Batch [{batch.id}]",
                    message=f"{user.name} submitted a bulk travel log {batch.file_name} for approval.",
                    type='info'
                )
                
                Notification.objects.create(
                    user=user,
                    title=f"Tour Plan Submitted [{batch.id}]",
                    message=f"Your Tour Plan (ID: {batch.id}) has been successfully submitted.",
                    type='success'
                )
            
            return Response(BulkActivityBatchSerializer(batch).data)
        except Exception as e:
            return Response({"error": f"Failed to parse Excel: {str(e)}"}, status=400)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        
        if batch.status not in ['Submitted', 'Manager Approved']:
            return Response({"error": "Batch is not in a valid status for approval"}, status=400)
            
        if batch.current_approver != user:
             return Response({"error": "Unauthorized: You are not the designated approver."}, status=403)
            
        is_hr = _is_hr(user)
        requester = batch.user
        
        # If the approver is not HR, do management chain logic
        if not is_hr:
            # Try explicit snapshots first (as like in trip approval)
            if batch.hierarchy_level == 1:
                next_approver = batch.senior_manager or batch.hod_director
            elif batch.hierarchy_level == 2:
                next_approver = batch.hod_director
            
            # DYNAMIC FALLBACK: If no snapshot level matches, try relative to current approver
            if not next_approver:
                potential_manager = user.reporting_manager
                if potential_manager and potential_manager != user and potential_manager != requester:
                    next_approver = potential_manager

            if next_approver:
                # Move to next manager in chain
                batch.current_approver = next_approver
                batch.hierarchy_level += 1
                batch.save()
                
                Notification.objects.create(
                    user=next_approver,
                    title=f"Pending Tour Plan Approval [{batch.id}]",
                    message=f"{requester.name}'s Tour Plan {batch.file_name} (ID: {batch.id}) requires your review (Forwarded from {user.name}).",
                    type='info'
                )
                Notification.objects.create(
                    user=requester,
                    title=f"Approved by {user.name} [{batch.id}]",
                    message=f"Your Tour Plan {batch.file_name} has been approved by {user.name} and forwarded to {next_approver.name} for review.",
                    type='success'
                )
                return Response({"message": f"Batch approved and forwarded to {next_approver.name}."})
            else:
                # End of Management Chain -> Move to HR Approval
                hr_head = get_hr_head(requester)
                batch.status = 'Manager Approved'
                batch.current_approver = hr_head
                batch.save()
                
                Notification.objects.create(
                    user=requester,
                    title=f"Management Approved [{batch.id}]",
                    message=f"Your Tour Plan {batch.file_name} has been approved by management and sent to HR for verification.",
                    type='success'
                )
                
                if hr_head:
                    Notification.objects.create(
                        user=hr_head,
                        title=f"HR Verification Required [{batch.id}]",
                        message=f"{requester.name}'s Tour Plan {batch.file_name} is management-approved and awaits your verification.",
                        type='info'
                    )
                return Response({"message": "Sent to HR for verification"})

        # --- STAGE 2: HR Approval ---
        # If we reach here, it's the final approval (HR)
        # --- Final Approval Side Effects ---
        # Activate Trip Story
        if batch.trip:
            update_trip_lifecycle(batch.trip, "Trip Story Activated", "Bulk activity log approved. Trip story and claims are now active.")
            # If there's a status that signifies 'Ready for Claim', we could set it here.
            # Usually 'Approved' or 'Completed' works.
            if batch.trip.consider_as_local:
                batch.trip.status = 'Approved' # Active for local
                batch.trip.save()
        # Create Expenses (Activities) for each row only when the final manager approves
        created_ids = []
        for row in batch.data_json:
            # 1. Get Origin and Destination from JSON
            origin = str(row.get('origin_route', '')).strip()
            destination = str(row.get('destination_route', '')).strip()

            # 2. Map mode and subType to match DynamicExpenseGrid's expected options
            # Since mode is not provided in the bulk template, we default to Car/Own Car
            excel_mode = str(row.get('mode', '')).lower()
            excel_vehicle = str(row.get('vehicle', '')).lower()
            
            mapped_mode = "Car / Cab"
            mapped_subType = "Own Car"
            
            if 'bike' in excel_mode or '2 wheeler' in excel_mode:
                mapped_mode = "Bike"
                mapped_subType = "Own Bike" if 'own' in excel_vehicle else "Ride Bike"
            elif 'bus' in excel_mode or 'metro' in excel_mode or 'public' in excel_mode:
                mapped_mode = "Public Transport"
                if 'metro' in excel_mode: mapped_subType = "Metro"
                elif 'bus' in excel_mode: mapped_subType = "Local Bus"
                else: mapped_subType = "Auto" # Default for PT
            elif excel_mode: # Only if provided
                if 'own' in excel_vehicle: mapped_subType = "Own Car"
                elif 'company' in excel_vehicle: mapped_subType = "Company Car"
                elif 'ride' in excel_vehicle or 'uber' in excel_vehicle or 'ola' in excel_vehicle or 'taxi' in excel_mode: 
                    mapped_subType = "Ride Hailing"

            # Pull and normalise the date. skip any bogus rows (e.g. the
            # instruction/sample line that sneaked in from earlier uploads).
            clean_date = str(row.get('date', '')).strip()
            if not clean_date or 'instruc' in clean_date.lower():
                # invalid/empty date – skip this row entirely
                continue
            # ensure format is YYYY-MM-DD; try parsing to detect bad values
            try:
                # this will raise ValueError for non‑ISO strings like "?  Instruc"
                datetime.date.fromisoformat(clean_date)
            except Exception:
                continue

            if len(clean_date) > 10:
                clean_date = clean_date[:10]

            desc_json = {
                "natureOfVisit": row.get('visit_intent'), # Updated to use visit_intent as purpose was removed
                "remarks": row.get('remarks'),
                "from_bulk_upload": True,
                "origin": origin,
                "destination": destination,
                "mode": mapped_mode,
                "subType": mapped_subType,
                "odoStart": row.get('odo_start', 0),
                "odoEnd": row.get('odo_end', 0),
                "time": {
                    "boardingDate": clean_date,
                    "boardingTime": "09:00",
                    "actualTime": "18:00"
                }
            }
            
            def safe_float(val):
                try: 
                    return float(val) if val not in [None, '', '-'] else 0.0
                except: 
                    return 0.0

            odo_s = safe_float(row.get('odo_start', 0))
            odo_e = safe_float(row.get('odo_end', 0))

            expense = Expense.objects.create(
                trip=batch.trip,
                date=clean_date,
                category='Fuel', 
                amount=0,       
                paid_by='Self (Out of Pocket)',
                description=json.dumps(desc_json),
                status='Approved', 
                odo_start=odo_s,
                odo_end=odo_e,
                distance=max(0, odo_e - odo_s),
                travel_mode=mapped_mode,
                vehicle_type=mapped_subType,
                latitude=0, longitude=0
            )
            created_ids.append(expense.id)
            
        batch.status = 'Approved'
        batch.created_expenses = created_ids
        batch.save()
        
        # Notify user
        Notification.objects.create(
            user=batch.user,
            title=f"Tour Plan Approved [{batch.id}]",
            message=f"Your Tour Plan {batch.file_name} has been final-approved and added to your report.",
            type='success'
        )
        
        return Response({"message": "Batch approved and activities created"})

    @action(detail=True, methods=['post'], url_path='reject')
    def reject_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        
        if batch.current_approver != user:
             return Response({"error": "Unauthorized"}, status=403)
             
        if 'data_json' in request.data and request.data['data_json'] is not None:
            batch.data_json = request.data['data_json']
             
        batch.status = 'Rejected'
        batch.remarks = request.data.get('remarks')
        if not batch.remarks:
            return Response({"error": "Rejection remarks are mandatory"}, status=400)
            
        batch.save()
        
        # Notify user
        Notification.objects.create(
            user=batch.user,
            title=f"Tour Plan Rejected [{batch.id}]",
            message=f"Your Tour Plan {batch.file_name} was rejected. Reason: {batch.remarks}",
            type='error'
        )
        return Response({"message": "Batch rejected"})

# --- MASTER VIEWSETS ---

class TravelModeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelModeMaster.objects.all()
    serializer_class = TravelModeMasterSerializer

class BookingTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = BookingTypeMaster.objects.all()
    serializer_class = BookingTypeMasterSerializer

class AirlineMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = AirlineMaster.objects.all()
    serializer_class = AirlineMasterSerializer

class TravelOperatorMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelOperatorMaster.objects.all()
    serializer_class = TravelOperatorMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_flight = self.request.query_params.get('is_flight')
        is_train = self.request.query_params.get('is_train')
        is_bus = self.request.query_params.get('is_bus')
        
        if is_flight is not None: queryset = queryset.filter(is_flight=is_flight.lower() == 'true')
        if is_train is not None: queryset = queryset.filter(is_train=is_train.lower() == 'true')
        if is_bus is not None: queryset = queryset.filter(is_bus=is_bus.lower() == 'true')
        return queryset

class TravelClassMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelClassMaster.objects.all()
    serializer_class = TravelClassMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_flight = self.request.query_params.get('is_flight')
        is_train = self.request.query_params.get('is_train')
        is_bus = self.request.query_params.get('is_bus')
        
        if is_flight is not None: queryset = queryset.filter(is_flight=is_flight.lower() == 'true')
        if is_train is not None: queryset = queryset.filter(is_train=is_train.lower() == 'true')
        if is_bus is not None: queryset = queryset.filter(is_bus=is_bus.lower() == 'true')
        return queryset

class BusTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = BusTypeMaster.objects.all()
    serializer_class = BusTypeMasterSerializer

class IntercityCabVehicleMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = IntercityCabVehicleMaster.objects.all()
    serializer_class = IntercityCabVehicleMasterSerializer

class TravelProviderMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelProviderMaster.objects.all()
    serializer_class = TravelProviderMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_flight = self.request.query_params.get('is_flight')
        is_train = self.request.query_params.get('is_train')
        is_bus = self.request.query_params.get('is_bus')
        is_intercity_cab = self.request.query_params.get('is_intercity_cab')
        
        if is_flight is not None: queryset = queryset.filter(is_flight=is_flight.lower() == 'true')
        if is_train is not None: queryset = queryset.filter(is_train=is_train.lower() == 'true')
        if is_bus is not None: queryset = queryset.filter(is_bus=is_bus.lower() == 'true')
        if is_intercity_cab is not None: queryset = queryset.filter(is_intercity_cab=is_intercity_cab.lower() == 'true')
        return queryset

class TravelVehicleMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelVehicleMaster.objects.all()
    serializer_class = TravelVehicleMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_bus = self.request.query_params.get('is_bus')
        is_intercity_cab = self.request.query_params.get('is_intercity_cab')
        
        if is_bus is not None: queryset = queryset.filter(is_bus=is_bus.lower() == 'true')
        if is_intercity_cab is not None: queryset = queryset.filter(is_intercity_cab=is_intercity_cab.lower() == 'true')
        return queryset


class JobReportViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = JobReport.objects.all()
    serializer_class = JobReportSerializer

    def get_queryset(self):
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            return self.queryset.filter(trip_id=trip_id)
        return self.queryset

    def perform_create(self, serializer):
        user = self.request.custom_user
        rm = user.reporting_manager
        sm = user.senior_manager
        hod = user.hod_director
        
        # Resolve initial approver
        current_approver = resolve_approver(user)
        
        serializer.save(
            current_approver=current_approver,
            status='Submitted',
            submitted_at=timezone.now(),
            reporting_manager=rm,
            senior_manager=sm,
            hod_director=hod,
            reporting_manager_name=rm.name if rm else None,
            senior_manager_name=sm.name if sm else None,
            hod_director_name=hod.name if hod else None
        )
        
        if current_approver:
             Notification.objects.create(
                user=current_approver,
                title="New Expense Claim",
                message=f"{user.name} submitted an expense claim for approval.",
                type='info'
            )

class LocalTravelModeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalTravelModeMaster.objects.all()
    serializer_class = LocalTravelModeMasterSerializer

class LocalSubTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalSubTypeMaster.objects.all()
    serializer_class = LocalSubTypeMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_car = self.request.query_params.get('is_car')
        is_bike = self.request.query_params.get('is_bike')
        is_auto = self.request.query_params.get('is_auto')
        
        if is_car is not None: queryset = queryset.filter(is_car=is_car.lower() == 'true')
        if is_bike is not None: queryset = queryset.filter(is_bike=is_bike.lower() == 'true')
        if is_auto is not None: queryset = queryset.filter(is_auto=is_auto.lower() == 'true')
        return queryset

class LocalProviderMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalProviderMaster.objects.all()
    serializer_class = LocalProviderMasterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        is_car = self.request.query_params.get('is_car')
        is_bike = self.request.query_params.get('is_bike')
        is_auto = self.request.query_params.get('is_auto')
        is_bus = self.request.query_params.get('is_bus')
        is_metro = self.request.query_params.get('is_metro')
        
        if is_car is not None: queryset = queryset.filter(is_car=is_car.lower() == 'true')
        if is_bike is not None: queryset = queryset.filter(is_bike=is_bike.lower() == 'true')
        if is_auto is not None: queryset = queryset.filter(is_auto=is_auto.lower() == 'true')
        if is_bus is not None: queryset = queryset.filter(is_bus=is_bus.lower() == 'true')
        if is_metro is not None: queryset = queryset.filter(is_metro=is_metro.lower() == 'true')
        return queryset

class StayTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayTypeMaster.objects.all()
    serializer_class = StayTypeMasterSerializer

class RoomTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = RoomTypeMaster.objects.all()
    serializer_class = RoomTypeMasterSerializer

class MealCategoryMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealCategoryMaster.objects.all()
    serializer_class = MealCategoryMasterSerializer

class MealTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealTypeMaster.objects.all()
    serializer_class = MealTypeMasterSerializer

class IncidentalTypeMasterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = IncidentalTypeMaster.objects.all()
    serializer_class = IncidentalTypeMasterSerializer

class MasterModuleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MasterModule.objects.all()
    serializer_class = MasterModuleSerializer

class CustomMasterDefinitionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterDefinition.objects.all()
    serializer_class = CustomMasterDefinitionSerializer

class CustomMasterValueViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterValue.objects.all()
    serializer_class = CustomMasterValueSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['definition']


